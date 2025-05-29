from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import requests, re, io
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import sys
import os

app = Flask(__name__)
CORS(app)
app.secret_key = "tu_clave_secreta"

def obtener_dolar_oficial():
    try:
        r = requests.get("https://www.bna.com.ar/", timeout=5)
        s = BeautifulSoup(r.text, "html.parser")
        tag = s.select_one("span.value.sell")
        return float(tag.text.strip().replace(".", "").replace(",", ".")) if tag else 1155.0
    except:
        return 1155.0

def scrapear(version, year):
    query = f"{version} {year}"
    url_base = "https://autos.mercadolibre.com.ar/" + query.replace(" ", "-")
    headers = {"User-Agent":"Mozilla/5.0"}
    rows, offset, page = [], 0, 1

    while True:
        url = url_base if page == 1 else f"{url_base}_Desde_{offset+1}"
        resp = requests.get(url, headers=headers)
        if resp.status_code != 200:
            break
        soup = BeautifulSoup(resp.text, "html.parser")
        items = soup.select("li.ui-search-layout__item")
        if not items:
            break

        for it in items:
            c = it.select_one("div.poly-card__content")
            if not c:
                continue

            # Título + link
            a = c.select_one("h3.poly-component__title-wrapper a")
            titulo = a.text.strip() if a else "N/D"
            link = a["href"] if a and a.has_attr("href") else "#"

            # Vendedor
            vend = c.select_one("span.poly-component__seller")
            vendedor = vend.text.strip() if vend else ""

            # Precio bruto + USD?
            raw = c.select_one("div.poly-component__price").get_text(" ", strip=True)
            m = re.search(r"([\d\.,]+)", raw)
            num = m.group(1).replace(".", "").replace(",", "") if m else "0"
            bruto = int(num) if num.isdigit() else 0
            es_usd = "U$S" in raw or "US$" in raw

            # Modelo y km
            attrs = c.select("div.poly-component__attributes-list ul.poly-attributes_list li")
            modelo = attrs[0].text.strip() if len(attrs) > 0 else ""
            km_str = attrs[1].text.strip() if len(attrs) > 1 else ""
            kmn = int(re.sub(r"[^\d]", "", km_str)) if km_str else 0

            # Ubicación
            loc = c.select_one("span.poly-component__location")
            ubic = loc.text.strip() if loc else ""

            rows.append({
                "titulo": titulo,
                "link": link,
                "vendedor": vendedor,
                "bruto": bruto,
                "es_usd": es_usd,
                "modelo": modelo,
                "km": kmn,
                "ubic": ubic
            })

        offset += len(items)
        page += 1

    # Conversiones y rangos
    dolar = obtener_dolar_oficial()
    precios = [(r["bruto"] * dolar if r["es_usd"] else r["bruto"]) for r in rows]
    kms = [r["km"] for r in rows]
    for i, r in enumerate(rows):
        r["precio_ars"] = precios[i]

    promedio = round(sum(precios) / len(precios)) if precios else 0
    price_min = min(precios) if precios else 0
    price_max = max(precios) if precios else 0
    km_min = min(kms) if kms else 0
    km_max = max(kms) if kms else 0

    return rows, promedio, price_min, price_max, km_min, km_max

@app.route("/api/search", methods=["POST"])
def api_search():
    data = request.get_json()
    version = data.get("version", "").strip()
    year = data.get("year", "2025")
    
    if not version:
        return jsonify({"error": "Versión requerida"}), 400
    
    try:
        datos, promedio, price_min, price_max, km_min, km_max = scrapear(version, year)
        return jsonify({
            "datos": datos,
            "promedio": promedio,
            "price_min": price_min,
            "price_max": price_max,
            "km_min": km_min,
            "km_max": km_max
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/export", methods=["POST"])
def api_export():
    data = request.get_json()
    version = data.get("version", "").strip()
    year = data.get("year", "2025")
    
    if not version:
        return jsonify({"error": "Versión requerida"}), 400
    
    try:
        rows, _, _, _, _, _ = scrapear(version, year)
        df = pd.DataFrame(rows)
        df_excel = pd.DataFrame({
            "Título": df["titulo"],
            "Precio (ARS)": [None]*len(df),
            "Vendedor": df["vendedor"],
            "Modelo": df["modelo"],
            "Kilómetros": df["km"],
            "Ubicación": df["ubic"],
            "Link": df["link"]
        })

        output = io.BytesIO()
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        link_font = Font(color="0000FF", underline="single")
        dolar = obtener_dolar_oficial()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_excel.to_excel(writer, index=False, sheet_name="Resultados")
            wb = writer.book
            ws = wb["Resultados"]
            last = len(rows) + 1

            for i, r in enumerate(rows, start=2):
                if r["es_usd"]:
                    cell = ws.cell(row=i, column=2, value=f"={r['bruto']}*{dolar}")
                    cell.fill = red
                else:
                    cell = ws.cell(row=i, column=2, value=r["bruto"])
                cell.number_format = u'"$"#,##0'

                link_cell = ws.cell(row=i, column=7, value="LINK a MeLi")
                link_cell.hyperlink = r["link"]
                link_cell.font = link_font

            ws.cell(row=last+1, column=1, value="PRECIO PROMEDIO")
            avgc = ws.cell(row=last+1, column=2, value=f"=AVERAGE(B2:B{last})")
            avgc.number_format = u'"$"#,##0'
            ws.auto_filter.ref = f"A1:G{last}"
            for col in range(1,8):
                ml = len(ws.cell(row=1,column=col).value)
                for ri in range(2,last+2):
                    v = ws.cell(row=ri,column=col).value
                    if v and len(str(v))>ml: ml = len(str(v))
                ws.column_dimensions[get_column_letter(col)].width = ml+2

        output.seek(0)
        filename = f"{version}-{year}-{datetime.now():%d-%m-%Y}.xlsx"
        return send_file(output,
                         download_name=filename,
                         as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/health", methods=["GET"])
def health_check():
    return jsonify({"status": "ok", "message": "Python API is running"})

if __name__=="__main__":
    print("Iniciando servidor Python en puerto 5001...")
    import sys
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    app.run(host="0.0.0.0", port=5001, debug=False)
